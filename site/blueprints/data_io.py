import io
from datetime import datetime, UTC, timedelta
from flask import Blueprint, jsonify, make_response, request, session
from utils.auth_decorators import login_required, admin_required
from werkzeug.utils import secure_filename
from utils.source_helpers import resolve_source_param, build_source_filter
from database import (
    get_db,
    create_scheduled_export,
    get_user_scheduled_exports,
    delete_scheduled_export,
    import_csv_readings,
    get_csv_import_stats,
    log_audit,
)


def _limit_decorator(limiter):
    if limiter:
        return limiter.limit

    def passthrough(*_args, **_kwargs):
        def wrapper(func):
            return func

        return wrapper

    return passthrough


def create_data_io_blueprint(limiter=None):
    limit = _limit_decorator(limiter)
    bp = Blueprint("data_io", __name__, url_prefix="/api")

    @bp.route("/export/json")
    @login_required
    @limit("10 per minute")
    def export_json():
        user_id = session.get("user_id")
        days = request.args.get("days", 30, type=int)
        db_source = resolve_source_param(allow_sim=False, allow_import=True)
        source_clause, source_params = build_source_filter(db_source)

        db = get_db()
        readings = db.execute(
            f"""
            SELECT timestamp, ppm FROM co2_readings
            WHERE timestamp >= datetime('now', '-' || ? || ' days')
            AND {source_clause}
            AND user_id = ?
            ORDER BY timestamp DESC
            """,
            (days, *source_params, user_id),
        ).fetchall()
        db.close()

        data = {
            "export_date": datetime.now(UTC).isoformat(),
            "days": days,
            "count": len(readings),
            "readings": [{"timestamp": r["timestamp"], "ppm": r["ppm"]} for r in readings],
        }

        return jsonify(data)

    @bp.route("/export/csv")
    @login_required
    @limit("10 per minute")
    def export_csv():
        user_id = session.get("user_id")
        days = request.args.get("days", 30, type=int)
        db_source = resolve_source_param(allow_sim=False, allow_import=True)
        source_clause, source_params = build_source_filter(db_source)

        db = get_db()
        readings = db.execute(
            f"""
            SELECT timestamp, ppm FROM co2_readings
            WHERE timestamp >= datetime('now', '-' || ? || ' days')
            AND {source_clause}
            AND user_id = ?
            ORDER BY timestamp DESC
            """,
            (days, *source_params, user_id),
        ).fetchall()
        db.close()

        csv_content = "timestamp,ppm\n" + "\n".join([f"{row['timestamp']},{row['ppm']}" for row in readings]) + "\n"
        response = make_response(csv_content)
        response.headers["Content-Disposition"] = f'attachment; filename="co2_export_{days}d.csv"'
        response.headers["Content-Type"] = "text/csv"
        return response

    @bp.route("/export/excel")
    @login_required
    @limit("10 per minute")
    def export_excel():
        user_id = session.get("user_id")
        days = request.args.get("days", 30, type=int)
        db_source = resolve_source_param(allow_sim=False, allow_import=True)
        source_clause, source_params = build_source_filter(db_source)

        db = get_db()
        readings = db.execute(
            f"""
            SELECT timestamp, ppm FROM co2_readings
            WHERE timestamp >= datetime('now', '-' || ? || ' days')
            AND {source_clause}
            AND user_id = ?
            ORDER BY timestamp DESC
            """,
            (days, *source_params, user_id),
        ).fetchall()
        db.close()

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CO₂ Data"

            ws["A1"] = "Timestamp"
            ws["B1"] = "PPM"
            for cell in ["A1", "B1"]:
                ws[cell].font = Font(bold=True, color="FFFFFF")
                ws[cell].fill = PatternFill(start_color="3DD98F", end_color="3DD98F", fill_type="solid")

            for idx, row in enumerate(readings, start=2):
                ws[f"A{idx}"] = row["timestamp"]
                ws[f"B{idx}"] = row["ppm"]

            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 15

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            response = make_response(output.getvalue())
            response.headers["Content-Disposition"] = f'attachment; filename="co2_export_{days}d.xlsx"'
            response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            return response
        except ImportError:
            return jsonify({"error": "Excel support not installed. Install openpyxl: pip install openpyxl"}), 400

    @bp.route("/export/schedule", methods=["POST"])
    @login_required
    def schedule_export():
        user_id = session.get("user_id")
        payload = request.get_json() or {}
        format = payload.get("format", "csv")
        frequency = payload.get("frequency", "weekly")

        if format not in ["csv", "json", "excel"]:
            return jsonify({"error": "Invalid format"}), 400

        if frequency not in ["daily", "weekly", "monthly"]:
            return jsonify({"error": "Invalid frequency"}), 400

        create_scheduled_export(user_id, format, frequency)
        return jsonify({"success": True, "format": format, "frequency": frequency})

    @bp.route("/export/scheduled")
    @login_required
    def get_scheduled_exports():
        user_id = session.get("user_id")
        exports = get_user_scheduled_exports(user_id)
        return jsonify({"exports": [dict(e) for e in exports]})

    @bp.route("/export/scheduled/<int:export_id>", methods=["DELETE"])
    @login_required
    def remove_scheduled_export(export_id):
        user_id = session.get("user_id")
        db = get_db()
        export = db.execute("SELECT user_id FROM scheduled_exports WHERE id = ?", (export_id,)).fetchone()
        db.close()

        if not export or export["user_id"] != user_id:
            return jsonify({"error": "Not found"}), 404

        delete_scheduled_export(export_id)
        return jsonify({"success": True})

    @bp.route("/import/csv", methods=["POST"])
    @login_required
    @limit("5 per minute")
    def import_csv():
        from utils.csv_validator import CSVValidator
        from utils.api_responses import error_response, success_response

        user_id = session.get("user_id")

        if "file" not in request.files:
            return error_response("No file uploaded", status_code=400, error_code="NO_FILE")

        file = request.files["file"]
        if not file or not file.filename:
            return error_response("No file selected", status_code=400, error_code="EMPTY_FILE")

        filename = secure_filename(file.filename)
        if not filename or not filename.lower().endswith(".csv"):
            return error_response("File must be CSV format", status_code=400, error_code="INVALID_FORMAT")

        try:
            validator = CSVValidator()
            is_valid, validation_result = validator.validate_file(file.stream)
            if not is_valid:
                return error_response("CSV validation failed", status_code=422, error_code="VALIDATION_FAILED", details=validation_result)

            if validation_result["valid_rows"]:
                result = import_csv_readings(validation_result["valid_rows"], user_id=user_id)
                log_audit(
                    user_id,
                    "CSV_IMPORT",
                    "data",
                    0,
                    "imported_count",
                    str(result.get("imported", 0)),
                    request.remote_addr,
                )

                return success_response(
                    data={"imported": result.get("imported", 0), "validation": validation_result},
                    message=f"Successfully imported {result.get('imported', 0)} readings",
                )

            return error_response(
                "No valid readings found in CSV",
                status_code=400,
                error_code="NO_VALID_DATA",
                details=validation_result,
            )
        except Exception as e:
            from app import logger

            logger.exception(f"CSV import error: {str(e)}")
            return error_response(f"Failed to process CSV: {str(e)}", status_code=500, error_code="PROCESSING_ERROR")

    @bp.route("/import/stats")
    @login_required
    @admin_required
    def import_stats():
        stats = get_csv_import_stats()
        return jsonify(stats)

    @bp.route("/export/simulate", methods=["POST"])
    @login_required
    def simulate_export():
        try:
            data = request.get_json() or {}
            export_format = data.get("format", "csv")
            period_days = data.get("period_days", 7)

            import random

            readings = []
            base_time = datetime.now(UTC) - timedelta(days=period_days)

            for i in range(period_days * 24):
                timestamp = base_time + timedelta(hours=i)
                co2_level = 800 + random.randint(-100, 300)
                readings.append({"timestamp": timestamp.isoformat(), "co2": co2_level})

            if export_format == "json":
                return jsonify(
                    {
                        "success": True,
                        "format": "json",
                        "export_date": datetime.now(UTC).isoformat(),
                        "period_days": period_days,
                        "records": len(readings),
                        "data": readings,
                    }
                )
            if export_format == "csv":
                csv_content = "timestamp,co2_ppm\n" + "\n".join([f"{r['timestamp']},{r['co2']}" for r in readings]) + "\n"
                response = make_response(csv_content)
                response.headers["Content-Disposition"] = f'attachment; filename="co2_export_{period_days}d.csv"'
                response.headers["Content-Type"] = "text/csv"
                return response
            return jsonify({"error": "Format non supporté"}), 400
        except Exception as e:
            return jsonify({"success": False, "error": str(e)}), 500

    return bp
